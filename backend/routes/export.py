"""Blueprint for exporting 立翔水電行 reports."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from flask import Blueprint, current_app, jsonify, redirect, send_file
from openpyxl import Workbook
from sqlalchemy.orm import selectinload

from decorators import role_required
from models import Task, TaskAssignee


export_bp = Blueprint("export", __name__)


def _serve_storage_file(filename: str):
    storage = current_app.extensions.get("storage")
    if storage is None:
        return jsonify({"msg": "Storage backend is not configured"}), 500

    if hasattr(storage, "local_path"):
        try:
            path = storage.local_path(filename)
        except FileNotFoundError:
            return jsonify({"msg": "File not found"}), 404
        return send_file(path, as_attachment=True)

    try:
        url = storage.url_for(filename)
    except Exception:
        return jsonify({"msg": "Unable to generate download link"}), 500
    return redirect(url)


def _attachment_summary(task: Task, kind: str) -> str:
    return "; ".join(
        attachment.original_name or attachment.file_path
        for attachment in task.attachments
        if attachment.file_type == kind
    )


@export_bp.get("/tasks")
@role_required("admin", "hq_staff", "site_supervisor")
def export_tasks():
    tasks = (
        Task.query.options(
            selectinload(Task.attachments),
            selectinload(Task.updates),
            selectinload(Task.assignee),
            selectinload(Task.assignees).selectinload(TaskAssignee.user),
            selectinload(Task.assigner),
        )
        .order_by(Task.created_at.desc())
        .all()
    )

    workbook = Workbook()
    tasks_sheet = workbook.active
    tasks_sheet.title = "Tasks"
    tasks_sheet.append(
        [
            "任務ID",
            "任務名稱",
            "地點",
            "狀態",
            "建立者",
            "負責人",
            "預計完成時間",
            "實際完成時間",
            "總工時(小時)",
            "圖片",
            "語音",
            "簽名",
        ]
    )

    for task in tasks:
        assigned_names = [
            assignment.user.username
            for assignment in task.assignees
            if assignment.user is not None
        ]
        if not assigned_names and task.assignee:
            assigned_names.append(task.assignee.username)

        tasks_sheet.append(
            [
                task.id,
                task.title,
                task.location,
                task.status,
                task.assigner.username if task.assigner else "",
                ", ".join(assigned_names),
                task.expected_time.isoformat() if task.expected_time else "",
                task.completed_at.isoformat() if task.completed_at else "",
                task.total_work_hours(),
                _attachment_summary(task, "image"),
                _attachment_summary(task, "audio"),
                _attachment_summary(task, "signature"),
            ]
        )

    attachments_sheet = workbook.create_sheet("Attachments")
    attachments_sheet.append(
        ["任務ID", "檔案類型", "原始檔名", "上傳時間", "備註", "下載連結"]
    )
    for task in tasks:
        for attachment in task.attachments:
            attachments_sheet.append(
                [
                    task.id,
                    attachment.file_type,
                    attachment.original_name,
                    attachment.uploaded_at.isoformat() if attachment.uploaded_at else "",
                    attachment.note or "",
                    attachment.to_dict().get("url"),
                ]
            )

    time_sheet = workbook.create_sheet("TimeEntries")
    time_sheet.append(
        ["任務ID", "使用者", "開始時間", "結束時間", "工時(小時)"]
    )
    for task in tasks:
        for update in task.updates:
            if update.start_time or update.end_time:
                time_sheet.append(
                    [
                        task.id,
                        update.author.username if update.author else update.user_id,
                        update.start_time.isoformat() if update.start_time else "",
                        update.end_time.isoformat() if update.end_time else "",
                        update.work_hours or 0.0,
                    ]
                )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"taskgo_tasks_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    relative_path = f"reports/{filename}"

    storage = current_app.extensions.get("storage")
    if storage is None:
        return jsonify({"msg": "Storage backend is not configured"}), 500

    storage.save(relative_path, stream)
    download_url = storage.url_for(relative_path)

    return jsonify({"url": download_url, "filename": filename})


@export_bp.get("/download/<path:filename>")
@role_required("admin", "hq_staff", "site_supervisor")
def download_export(filename: str):
    return _serve_storage_file(f"reports/{filename}")

