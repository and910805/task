from __future__ import annotations

from datetime import date, datetime, time, timedelta
from io import BytesIO
import glob
import json
import os
from pathlib import Path
import re

from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import jwt_required
from openpyxl import load_workbook
from sqlalchemy.orm import selectinload

from decorators import role_required
from extensions import db
from models import Contact, Customer, Invoice, InvoiceItem, Quote, QuoteItem, QuoteVersion, ServiceCatalogItem, Task
from utils import get_current_user_id

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

crm_bp = Blueprint("crm", __name__)

READ_ROLES = ("site_supervisor", "hq_staff", "admin")
WRITE_ROLES = ("site_supervisor", "hq_staff", "admin")
VALID_QUOTE_STATUS = {"draft", "sent", "accepted", "rejected", "expired"}
VALID_INVOICE_STATUS = {"draft", "issued", "partially_paid", "paid", "cancelled"}

PDF_FONT_NAME = "Helvetica"
PDF_FONT_ENV = "PDF_FONT_PATH"
PDF_FONT_CANDIDATES = (
    "/usr/local/share/fonts/NotoSansTC-wght.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/noto/NotoSerifCJK-Regular.ttc",
    "/usr/share/fonts/opentype/source-han-sans/SourceHanSansTW-Regular.otf",
)
PDF_CID_FALLBACKS = ("MSung-Light", "STSong-Light")
PDF_REQUIRE_EMBEDDED_FONT_ENV = "PDF_REQUIRE_EMBEDDED_FONT"
# Use unicode escapes to avoid source-file encoding issues on Windows/editors.
PDF_CJK_PROBE = "\u4f30\u50f9\u55ae\u767c\u7968\u53f0\u7167"  # ???????
PDF_FONT_SOURCE = "default"
PDF_FONT_PATH_USED = ""
PDF_STAMP_ENV = "PDF_STAMP_IMAGE_PATH"
PDF_STAMP_DEFAULT_FILENAME = "S__5505135-removebg-preview.png"
PDF_STAMP_ROTATE_ENV = "PDF_STAMP_ROTATE_DEG"
PDF_STAMP_DEFAULT_ROTATE_DEG = 90.0
PDF_STAMP_WIDTH_MM = 24.0
PDF_STAMP_Y_OFFSET_ENV = "PDF_STAMP_Y_OFFSET_MM"
PDF_STAMP_DEFAULT_Y_OFFSET_MM = 10.0
FINANCIAL_DIGITS = ("零", "壹", "貳", "參", "肆", "伍", "陸", "柒", "捌", "玖")
FINANCIAL_SMALL_UNITS = ("", "拾", "佰", "仟")
FINANCIAL_BIG_UNITS = ("", "萬", "億", "兆")


def _font_supports_traditional_chinese(font_name: str) -> bool:
    if font_name in PDF_CID_FALLBACKS:
        return True
    try:
        font = pdfmetrics.getFont(font_name)
    except Exception:
        return False

    face = getattr(font, "face", None)
    char_widths = getattr(face, "charWidths", None)
    if isinstance(char_widths, dict) and all(ord(ch) in char_widths for ch in PDF_CJK_PROBE):
        return True

    # Some reportlab/font combinations don't expose complete charWidths for TTC fonts.
    # Fallback to a render-width probe to avoid false negatives in container deployments.
    try:
        width = float(pdfmetrics.stringWidth(PDF_CJK_PROBE, font_name, 12))
        return width > 0
    except Exception:
        return False


def _env_flag(name: str, default: bool = False) -> bool:
    raw = (os.environ.get(name) or "").strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on"}


def _candidate_pdf_font_paths() -> list[str]:
    paths: list[str] = []
    seen: set[str] = set()

    def _add(path: str) -> None:
        normalized = (path or "").strip()
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        paths.append(normalized)

    configured = os.environ.get(PDF_FONT_ENV, "").strip()
    if configured and os.path.exists(configured):
        _add(configured)

    for candidate in PDF_FONT_CANDIDATES:
        if os.path.exists(candidate):
            _add(candidate)

    # Runtime fallback for distros where the package path differs.
    dynamic_patterns = (
        "/usr/local/share/fonts/**/*NotoSansTC*.ttf",
        "/usr/local/share/fonts/**/*NotoSerifTC*.ttf",
        "/usr/share/fonts/**/*NotoSansCJK*Regular*.ttc",
        "/usr/share/fonts/**/*NotoSerifCJK*Regular*.ttc",
        "/usr/share/fonts/**/*SourceHanSansTW*Regular*.otf",
        "/usr/share/fonts/**/*SourceHanSans*Regular*.otf",
    )
    for pattern in dynamic_patterns:
        for path in sorted(glob.glob(pattern, recursive=True)):
            if os.path.isfile(path):
                _add(path)
    return paths


def _discover_pdf_font_path() -> str:
    return (_candidate_pdf_font_paths() or [""])[0]


def _ensure_pdf_font():
    global PDF_FONT_NAME, PDF_FONT_SOURCE, PDF_FONT_PATH_USED

    if PDF_FONT_NAME in ("CustomFont", *PDF_CID_FALLBACKS) and _font_supports_traditional_chinese(PDF_FONT_NAME):
        return

    for font_path in _candidate_pdf_font_paths():
        if not font_path or not os.path.exists(font_path):
            continue
        # Noto CJK on Debian is usually a .ttc. Try multiple subfont indexes.
        # Prefer TC/HK indices first, then SC/KR/JP.
        ttc_indices = (3, 4, 2, 1, 0) if font_path.lower().endswith(".ttc") else (0,)
        for idx in ttc_indices:
            try:
                pdfmetrics.registerFont(TTFont("CustomFont", font_path, subfontIndex=idx))
                if _font_supports_traditional_chinese("CustomFont"):
                    PDF_FONT_NAME = "CustomFont"
                    PDF_FONT_SOURCE = "filesystem"
                    PDF_FONT_PATH_USED = font_path
                    return
            except Exception:
                continue

    # Optional fallback: built-in CID fonts are not embedded and may render incorrectly
    # on some mobile/desktop PDF readers. Keep this disabled by default.
    if not _env_flag(PDF_REQUIRE_EMBEDDED_FONT_ENV, default=True):
        for cid_name in PDF_CID_FALLBACKS:
            try:
                pdfmetrics.registerFont(UnicodeCIDFont(cid_name))
                if _font_supports_traditional_chinese(cid_name):
                    PDF_FONT_NAME = cid_name
                    PDF_FONT_SOURCE = f"cid:{cid_name}"
                    PDF_FONT_PATH_USED = ""
                    return
            except Exception:
                continue

    PDF_FONT_NAME = "Helvetica"
    PDF_FONT_SOURCE = "default"
    PDF_FONT_PATH_USED = ""


def _require_embedded_pdf_font() -> None:
    _ensure_pdf_font()
    if PDF_FONT_SOURCE == "filesystem" and _font_supports_traditional_chinese(PDF_FONT_NAME):
        return
    raise RuntimeError(
        "PDF font is not embedded-capable for Traditional Chinese. "
        f"Current source={PDF_FONT_SOURCE!r}. Set {PDF_FONT_ENV} to a CJK font file "
        "(e.g. NotoSansCJK-Regular.ttc)."
    )


def _pdf_font_health_payload() -> dict:
    _ensure_pdf_font()
    require_embedded = _env_flag(PDF_REQUIRE_EMBEDDED_FONT_ENV, default=True)
    embedded_ok = PDF_FONT_SOURCE == "filesystem" and _font_supports_traditional_chinese(PDF_FONT_NAME)
    configured_font_path = (os.environ.get(PDF_FONT_ENV) or "").strip()
    return {
        "font_name": PDF_FONT_NAME,
        "font_source": PDF_FONT_SOURCE,
        "font_path": PDF_FONT_PATH_USED,
        "configured_font_path": configured_font_path or None,
        "configured_font_path_exists": bool(configured_font_path and os.path.exists(configured_font_path)),
        "discovered_font_path": _discover_pdf_font_path() or None,
        "font_candidates_found": _candidate_pdf_font_paths(),
        "supports_traditional_chinese": _font_supports_traditional_chinese(PDF_FONT_NAME),
        "require_embedded_font": require_embedded,
        "embedded_font_ready": embedded_ok,
        "pdf_generation_ready": embedded_ok or not require_embedded,
        "probe_text": PDF_CJK_PROBE,
        "candidates": list(PDF_FONT_CANDIDATES),
        "hint": f"Set {PDF_FONT_ENV} to a Noto/SourceHan CJK font file path if PDF is not ready.",
    }


def _resolve_pdf_stamp_path() -> str | None:
    configured = (os.environ.get(PDF_STAMP_ENV) or "").strip()
    if configured and os.path.exists(configured):
        return configured

    backend_dir = Path(__file__).resolve().parents[1]
    default_path = backend_dir.parent / "data" / PDF_STAMP_DEFAULT_FILENAME
    if default_path.exists() and default_path.is_file():
        return str(default_path)
    return None


def _resolve_pdf_stamp_rotation_deg() -> float:
    raw = (os.environ.get(PDF_STAMP_ROTATE_ENV) or "").strip()
    if not raw:
        return PDF_STAMP_DEFAULT_ROTATE_DEG
    try:
        return float(raw)
    except ValueError:
        return PDF_STAMP_DEFAULT_ROTATE_DEG


def _resolve_pdf_stamp_y_offset_mm() -> float:
    raw = (os.environ.get(PDF_STAMP_Y_OFFSET_ENV) or "").strip()
    if not raw:
        return PDF_STAMP_DEFAULT_Y_OFFSET_MM
    try:
        value = float(raw)
        return max(-20.0, min(20.0, value))
    except ValueError:
        return PDF_STAMP_DEFAULT_Y_OFFSET_MM


def _flowable_render_height(flowable, avail_width: float, avail_height: float) -> float:
    width = max(avail_width, 1.0)
    height = max(avail_height, 1.0)
    _, wrapped_h = flowable.wrap(width, height)
    before = float(flowable.getSpaceBefore()) if hasattr(flowable, "getSpaceBefore") else 0.0
    after = float(flowable.getSpaceAfter()) if hasattr(flowable, "getSpaceAfter") else 0.0
    return float(wrapped_h) + before + after


def _estimate_table_cell_center(doc, flowables_before, table, row_index: int, col_index: int, h_align: str):
    try:
        table.wrap(doc.width, doc.height)
        row_heights = [float(v) for v in getattr(table, "_rowHeights", [])]
        col_widths = [float(v) for v in getattr(table, "_colWidths", [])]
        if not row_heights or not col_widths:
            return None
        if row_index < 0 or row_index >= len(row_heights):
            return None
        if col_index < 0 or col_index >= len(col_widths):
            return None

        used_height = 0.0
        remaining_height = float(doc.height)
        for flowable in flowables_before:
            block_h = _flowable_render_height(flowable, float(doc.width), remaining_height)
            used_height += block_h
            remaining_height = max(1.0, remaining_height - block_h)

        page_w, page_h = doc.pagesize
        table_w = sum(col_widths)
        align = (h_align or "LEFT").upper()
        if align == "RIGHT":
            table_left = float(doc.leftMargin) + float(doc.width) - table_w
        elif align == "CENTER":
            table_left = float(doc.leftMargin) + (float(doc.width) - table_w) / 2.0
        else:
            table_left = float(doc.leftMargin)

        table_top = float(page_h) - float(doc.topMargin) - used_height
        row_top = table_top - sum(row_heights[:row_index])
        row_center_y = row_top - row_heights[row_index] / 2.0
        col_left = table_left + sum(col_widths[:col_index])
        col_center_x = col_left + col_widths[col_index] / 2.0
        return col_center_x, row_center_y
    except Exception:
        return None


def _draw_pdf_stamp(canvas, doc, center: tuple[float, float] | None = None):
    stamp_path = _resolve_pdf_stamp_path()
    if not stamp_path:
        return
    try:
        image = ImageReader(stamp_path)
        src_w, src_h = image.getSize()
        if not src_w or not src_h:
            return

        stamp_w = PDF_STAMP_WIDTH_MM * mm
        stamp_h = stamp_w * float(src_h) / float(src_w)
        rotate_deg = _resolve_pdf_stamp_rotation_deg()
        y_offset = _resolve_pdf_stamp_y_offset_mm() * mm

        if center is not None:
            center_x, center_y = center
        else:
            page_w, page_h = doc.pagesize
            center_x = page_w - doc.rightMargin - (stamp_w / 2.0)
            center_y = page_h - doc.topMargin - (stamp_h / 2.0) + 4 * mm
        center_y += y_offset
        page_w, page_h = doc.pagesize
        min_x = float(doc.leftMargin) + (stamp_w / 2.0)
        max_x = float(page_w) - float(doc.rightMargin) - (stamp_w / 2.0)
        min_y = float(doc.bottomMargin) + (stamp_h / 2.0)
        max_y = float(page_h) - float(doc.topMargin) - (stamp_h / 2.0)
        center_x = max(min_x, min(max_x, float(center_x)))
        center_y = max(min_y, min(max_y, float(center_y)))

        canvas.saveState()
        canvas.translate(center_x, center_y)
        canvas.rotate(rotate_deg)
        canvas.drawImage(
            image,
            -stamp_w / 2.0,
            -stamp_h / 2.0,
            width=stamp_w,
            height=stamp_h,
            preserveAspectRatio=True,
            mask="auto",
        )
        canvas.restoreState()
    except Exception:
        return


def _make_pdf_stamp_canvasmaker(doc, center: tuple[float, float] | None = None):
    class _StampCanvas(pdf_canvas.Canvas):
        def showPage(self):
            _draw_pdf_stamp(self, doc, center)
            super().showPage()

    return _StampCanvas


def _parse_date(raw, field_name: str):
    if raw in (None, ""):
        return None, None
    if isinstance(raw, date):
        return raw, None
    try:
        return date.fromisoformat(str(raw)), None
    except ValueError:
        return None, (jsonify({"msg": f"Invalid {field_name} format, expected YYYY-MM-DD"}), 400)


def _parse_float(raw, field_name: str, *, minimum: float | None = None):
    if raw in (None, ""):
        return None, None
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return None, (jsonify({"msg": f"{field_name} must be a number"}), 400)
    if minimum is not None and value < minimum:
        return None, (jsonify({"msg": f"{field_name} must be >= {minimum}"}), 400)
    return value, None


def _normalize_items(raw_items):
    if not isinstance(raw_items, list) or not raw_items:
        return None, (jsonify({"msg": "items is required and must be a non-empty array"}), 400)

    normalized = []
    for idx, raw in enumerate(raw_items):
        if not isinstance(raw, dict):
            return None, (jsonify({"msg": f"items[{idx}] must be an object"}), 400)

        description = (raw.get("description") or "").strip()
        if not description:
            return None, (jsonify({"msg": f"items[{idx}].description is required"}), 400)
        unit = (raw.get("unit") or "").strip() or "式"

        qty, qty_err = _parse_float(raw.get("quantity", 1), f"items[{idx}].quantity", minimum=0)
        if qty_err:
            return None, qty_err
        unit_price_val, price_err = _parse_float(raw.get("unit_price", 0), f"items[{idx}].unit_price", minimum=0)
        if price_err:
            return None, price_err

        quantity = qty if qty is not None else 1.0
        unit_price = unit_price_val if unit_price_val is not None else 0.0
        amount = round(quantity * unit_price, 2)
        normalized.append(
            {
                "description": description,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "amount": amount,
                "sort_order": idx,
            }
        )

    return normalized, None


def _apply_totals(entity, items: list[dict], tax_rate_raw):
    subtotal = round(sum(item["amount"] for item in items), 2)
    tax_rate, tax_rate_err = _parse_float(tax_rate_raw, "tax_rate", minimum=0)
    if tax_rate_err:
        return tax_rate_err

    safe_tax_rate = round(tax_rate or 0.0, 2)
    tax_amount = round(subtotal * safe_tax_rate / 100.0, 2)
    total_amount = round(subtotal + tax_amount, 2)

    entity.subtotal = subtotal
    entity.tax_rate = safe_tax_rate
    entity.tax_amount = tax_amount
    entity.total_amount = total_amount
    return None


def _next_quote_no() -> str:
    ymd = datetime.utcnow().strftime("%Y%m%d")
    prefix = f"QT-{ymd}-"

    rows = Quote.query.with_entities(Quote.quote_no).filter(Quote.quote_no.like(f"{prefix}%")).all()
    max_seq = 0
    for (quote_no,) in rows:
        if not isinstance(quote_no, str) or not quote_no.startswith(prefix):
            continue
        suffix = quote_no[len(prefix):]
        if suffix.isdigit():
            max_seq = max(max_seq, int(suffix))

    next_seq = max_seq + 1
    candidate = f"{prefix}{next_seq:03d}"
    while Quote.query.filter(Quote.quote_no == candidate).first() is not None:
        next_seq += 1
        candidate = f"{prefix}{next_seq:03d}"
    return candidate


def _next_quote_version_no(quote_id: int) -> int:
    latest = (
        QuoteVersion.query.with_entities(QuoteVersion.version_no)
        .filter(QuoteVersion.quote_id == quote_id)
        .order_by(QuoteVersion.version_no.desc())
        .first()
    )
    return int(latest[0]) + 1 if latest else 1


def _append_quote_version_snapshot(quote: Quote, *, action: str, summary: str | None = None) -> None:
    quote_payload = quote.to_dict()
    db.session.add(
        QuoteVersion(
            quote_id=quote.id,
            version_no=_next_quote_version_no(quote.id),
            action=(action or "update").strip().lower() or "update",
            summary=(summary or "").strip() or None,
            snapshot_json=json.dumps(quote_payload, ensure_ascii=False),
            changed_by_id=get_current_user_id(),
        )
    )


def _invoice_module_disabled():
    return jsonify({"msg": "Invoice module is disabled in quote-only mode"}), 410


def _quote_display_total_without_tax(quote: Quote) -> float:
    # The template-style quote output is tax-exclusive.
    if quote.subtotal is not None:
        return float(quote.subtotal)
    return float(quote.total_amount or 0)


def _invoice_display_total_without_tax(invoice: Invoice) -> float:
    # Keep invoice PDF aligned with quote output: tax-exclusive amount.
    if invoice.subtotal is not None:
        return float(invoice.subtotal)
    return float(invoice.total_amount or 0)


def _to_roc_date_text(value: date | None) -> str:
    value = value or date.today()
    roc_year = value.year - 1911
    return f"中華民國  {roc_year} 年 {value.month} 月 {value.day} 日"


def _format_amount_number(amount: float) -> str:
    safe_amount = round(float(amount or 0), 2)
    if safe_amount.is_integer():
        return f"{safe_amount:,.0f}"
    return f"{safe_amount:,.2f}"


def _financial_group_to_text(group_value: int) -> str:
    if group_value <= 0:
        return ""
    text = ""
    zero_pending = False
    for unit_idx in range(3, -1, -1):
        base = 10**unit_idx
        digit = (group_value // base) % 10
        if digit == 0:
            if text:
                zero_pending = True
            continue
        if zero_pending:
            text += FINANCIAL_DIGITS[0]
            zero_pending = False
        text += FINANCIAL_DIGITS[digit] + FINANCIAL_SMALL_UNITS[unit_idx]
    return text


def _financial_integer_to_text(value: int) -> str:
    if value <= 0:
        return FINANCIAL_DIGITS[0]

    groups: list[int] = []
    while value > 0:
        groups.append(value % 10000)
        value //= 10000

    result: list[str] = []
    zero_between_groups = False
    for idx in range(len(groups) - 1, -1, -1):
        group_value = groups[idx]
        if group_value == 0:
            zero_between_groups = True
            continue
        if result and (zero_between_groups or group_value < 1000):
            result.append(FINANCIAL_DIGITS[0])
        zero_between_groups = False
        result.append(_financial_group_to_text(group_value))
        big_unit = FINANCIAL_BIG_UNITS[idx] if idx < len(FINANCIAL_BIG_UNITS) else ""
        if big_unit:
            result.append(big_unit)

    return "".join(result) if result else FINANCIAL_DIGITS[0]


def _format_financial_amount_text(amount: float) -> str:
    rounded = round(float(amount or 0), 2)
    integer_amount = int(round(rounded))
    return f"{_financial_integer_to_text(integer_amount)}元整"


def _default_quote_dates(issue_date: date | None, expiry_date: date | None) -> tuple[date, date]:
    safe_issue = issue_date or date.today()
    safe_expiry = expiry_date or (safe_issue + timedelta(days=10))
    return safe_issue, safe_expiry


def _resolve_quote_recipient_display(quote: Quote, customer: Customer | None, contact: Contact | None) -> str:
    customer_name = (customer.name if customer else "") or ""
    if customer_name.strip():
        return customer_name.strip()
    if quote.recipient_name and quote.recipient_name.strip():
        return quote.recipient_name.strip()
    contact_name = (contact.name if contact else "") or ""
    return contact_name.strip()


def _safe_download_filename_part(raw: str | None, fallback: str = "估價單") -> str:
    source = (raw or "").strip() or fallback
    safe = re.sub(r'[\\/:*?"<>|]+', "_", source)
    safe = re.sub(r"\s+", "_", safe).strip("._")
    return safe[:80] or fallback


def _quote_task_description(quote: Quote, customer: Customer | None, contact: Contact | None) -> str:
    ordered_items = sorted(
        quote.items,
        key=lambda item: (
            item.sort_order if item.sort_order is not None else 10**9,
            item.id if item.id is not None else 10**9,
        ),
    )
    lines = [
        f"由報價單自動建立任務：{quote.quote_no}",
        f"客戶：{(customer.name if customer else '') or '-'}",
        f"聯絡人：{(contact.name if contact else '') or '-'}",
        f"報價日期：{quote.issue_date.isoformat() if quote.issue_date else '-'}",
        f"有效日期：{quote.expiry_date.isoformat() if quote.expiry_date else '-'}",
        "",
        "品項摘要：",
    ]
    for idx, item in enumerate(ordered_items[:6], start=1):
        lines.append(
            f"{idx}. {item.description or '-'} | {float(item.quantity or 0):.2f}{item.unit or ''}"
        )
    if quote.note:
        lines.extend(["", f"備註：{quote.note}"])
    return "\n".join(lines)


def _create_task_for_quote(quote: Quote, customer: Customer | None, contact: Contact | None) -> Task:
    issue_value = quote.issue_date or date.today()
    expiry_value = quote.expiry_date or (issue_value + timedelta(days=10))
    title_name = (customer.name if customer else "") or "未命名客戶"
    title = f"報價單 {quote.quote_no} - {title_name}"[:150]
    location = ((customer.address if customer else "") or "待確認地址").strip()[:255] or "待確認地址"
    expected_time = datetime.combine(issue_value, time(hour=9, minute=0))
    due_date = datetime.combine(expiry_value, time(hour=18, minute=0))

    return Task(
        title=title,
        description=_quote_task_description(quote, customer, contact),
        status="尚未接單",
        location=location,
        location_url=None,
        expected_time=expected_time,
        assigned_to_id=None,
        assigned_by_id=quote.created_by_id,
        due_date=due_date,
    )


def _find_quote_template_path() -> Path | None:
    configured = (os.environ.get("QUOTE_TEMPLATE_XLSX") or "").strip()
    if configured:
        candidate = Path(configured)
        if candidate.exists() and candidate.is_file():
            return candidate

    backend_dir = Path(__file__).resolve().parents[1]
    data_dir = backend_dir.parent / "data"
    if not data_dir.exists():
        return None

    candidates = sorted(data_dir.glob("*.xlsx"))
    return candidates[0] if candidates else None


def _apply_quote_to_template_sheet(ws, quote: Quote, customer: Customer | None, contact: Contact | None) -> None:
    recipient = _resolve_quote_recipient_display(quote, customer, contact)

    ws["D2"] = "立翔水電工程行"
    ws["D3"] = "估價單"
    ws["D4"] = recipient
    ws["E4"] = "台照"
    ws["D5"] = _to_roc_date_text(quote.issue_date)

    # Template reserves rows 7-26 for up to 20 line items.
    for idx, row in enumerate(range(7, 27), start=1):
        ws[f"C{row}"] = idx
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None
        ws[f"F{row}"] = None
        ws[f"G{row}"] = None
        ws[f"H{row}"] = None
        ws[f"I{row}"] = 0
        ws[f"J{row}"] = None

    ordered_items = sorted(
        quote.items,
        key=lambda item: (
            item.sort_order if item.sort_order is not None else 10**9,
            item.id if item.id is not None else 10**9,
        ),
    )
    for index, item in enumerate(ordered_items[:20]):
        row = 7 + index
        ws[f"D{row}"] = item.description or ""
        ws[f"F{row}"] = item.unit or "式"
        ws[f"G{row}"] = float(item.quantity or 0)
        ws[f"H{row}"] = float(item.unit_price or 0)
        ws[f"I{row}"] = float(item.amount or 0)

    total_amount = _quote_display_total_without_tax(quote)
    ws["C27"] = "合計"
    ws["E27"] = "新台幣"
    ws["F27"] = total_amount
    ws["H27"] = "NT$"
    ws["I27"] = total_amount


def _validate_customer_contact(customer_id, contact_id):
    customer = Customer.query.get(customer_id)
    if not customer:
        return None, None, (jsonify({"msg": "Customer not found"}), 404)

    contact = None
    if contact_id is not None:
        contact = Contact.query.get(contact_id)
        if not contact:
            return None, None, (jsonify({"msg": "Contact not found"}), 404)
        if contact.customer_id != customer.id:
            return None, None, (jsonify({"msg": "Contact does not belong to this customer"}), 400)

    return customer, contact, None


def _serialize_customer_service_history(customer: Customer, *, quote_limit: int = 30, invoice_limit: int = 30):
    quotes = (
        Quote.query.options(selectinload(Quote.items))
        .filter(Quote.customer_id == customer.id)
        .order_by(Quote.created_at.desc())
        .limit(quote_limit)
        .all()
    )
    return {
        "customer": customer.to_dict(),
        "quotes": [row.to_dict() for row in quotes],
        "invoices": [],
    }


def _build_pdf_document(title: str, meta_rows: list[list[str]], item_rows: list[list[str]], totals_rows: list[list[str]]):
    _require_embedded_pdf_font()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = PDF_FONT_NAME
    styles["Heading1"].fontName = PDF_FONT_NAME

    story = [
        Paragraph("立翔水電行", styles["Heading1"]),
        Paragraph(title, styles["Normal"]),
        Spacer(1, 8 * mm),
    ]

    meta_table = Table(meta_rows, hAlign="LEFT", colWidths=[45 * mm, 120 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
            ]
        )
    )
    story.extend([meta_table, Spacer(1, 6 * mm)])

    item_col_count = len(item_rows[0]) if item_rows else 0
    if item_col_count == 5:
        col_widths = [72 * mm, 18 * mm, 20 * mm, 30 * mm, 30 * mm]
    else:
        col_widths = [80 * mm, 25 * mm, 30 * mm, 30 * mm]
    items_table = Table(item_rows, hAlign="LEFT", colWidths=col_widths)
    items_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5f5")),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    story.extend([items_table, Spacer(1, 6 * mm)])

    totals_table = Table(totals_rows, hAlign="RIGHT", colWidths=[35 * mm, 35 * mm])
    totals_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
            ]
        )
    )
    totals_row_index = len(totals_rows) - 1 if totals_rows else 0
    stamp_center = _estimate_table_cell_center(
        doc,
        story,
        totals_table,
        row_index=totals_row_index,
        col_index=1,
        h_align="RIGHT",
    )
    story.append(totals_table)

    doc.build(story, canvasmaker=_make_pdf_stamp_canvasmaker(doc, stamp_center))
    buffer.seek(0)
    return buffer


def _build_quote_template_pdf(quote: Quote, customer: Customer | None, contact: Contact | None):
    _require_embedded_pdf_font()
    recipient = _resolve_quote_recipient_display(quote, customer, contact)

    ordered_items = sorted(
        quote.items,
        key=lambda item: (
            item.sort_order if item.sort_order is not None else 10**9,
            item.id if item.id is not None else 10**9,
        ),
    )

    rows = [["項目", "項目名稱", "規格內容", "單位", "數量", "單價", "總額", "備註"]]
    for idx in range(20):
        item = ordered_items[idx] if idx < len(ordered_items) else None
        if item is None:
            rows.append([str(idx + 1), "", "", "", "", "", "", ""])
            continue
        rows.append(
            [
                str(idx + 1),
                item.description or "",
                "",
                item.unit or "式",
                f"{float(item.quantity or 0):.2f}",
                f"{float(item.unit_price or 0):.2f}",
                f"{float(item.amount or 0):.2f}",
                "",
            ]
        )

    total_amount = _quote_display_total_without_tax(quote)
    total_amount_numeric = _format_amount_number(total_amount)
    total_amount_upper = _format_financial_amount_text(total_amount)
    rows.append(
        [
            "合計",
            "",
            "新台幣",
            total_amount_upper,
            "",
            "NT$",
            total_amount_numeric,
            "",
        ]
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"quote-{quote.quote_no}",
    )

    styles = getSampleStyleSheet()
    title_style = styles["Heading1"].clone("QuoteTemplateTitle")
    title_style.fontName = PDF_FONT_NAME
    title_style.alignment = 1
    title_style.fontSize = 24
    title_style.leading = 30
    title_style.textColor = colors.HexColor("#111827")

    subtitle_style = styles["Normal"].clone("QuoteTemplateSubtitle")
    subtitle_style.fontName = PDF_FONT_NAME
    subtitle_style.alignment = 1
    subtitle_style.fontSize = 13
    subtitle_style.leading = 18
    subtitle_style.textColor = colors.HexColor("#334155")

    body_style = styles["Normal"].clone("QuoteTemplateBody")
    body_style.fontName = PDF_FONT_NAME
    body_style.fontSize = 10
    body_style.leading = 14
    body_style.textColor = colors.HexColor("#1f2937")
    signer_style = styles["Normal"].clone("QuoteTemplateSigner")
    signer_style.fontName = PDF_FONT_NAME
    signer_style.alignment = 2
    signer_style.fontSize = 12
    signer_style.leading = 16

    story = [
        Paragraph("立翔水電工程行", title_style),
        Paragraph("估價單", subtitle_style),
        Spacer(1, 3 * mm),
        Paragraph(f"{recipient} 台照", body_style),
        Paragraph(_to_roc_date_text(quote.issue_date), body_style),
        Spacer(1, 4 * mm),
    ]

    table = Table(
        rows,
        colWidths=[12 * mm, 46 * mm, 28 * mm, 14 * mm, 14 * mm, 20 * mm, 20 * mm, 20 * mm],
        repeatRows=1,
        hAlign="CENTER",
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (5, -1), "CENTER"),
                ("ALIGN", (6, 0), (6, -1), "RIGHT"),
                ("ALIGN", (7, 0), (7, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#9ca3af")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#fafafa")]),
                ("BACKGROUND", (6, 1), (6, -1), colors.HexColor("#fff3a3")),
                ("FONTNAME", (0, -1), (-1, -1), PDF_FONT_NAME),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2ff")),
                ("LINEABOVE", (0, -1), (-1, -1), 1.0, colors.HexColor("#4b5563")),
                ("SPAN", (3, -1), (4, -1)),
                ("ALIGN", (3, -1), (4, -1), "LEFT"),
                ("FONTSIZE", (3, -1), (4, -1), 11),
                ("FONTSIZE", (6, -1), (6, -1), 11),
            ]
        )
    )
    totals_row_index = len(rows) - 1 if rows else 0
    stamp_center = _estimate_table_cell_center(
        doc,
        story,
        table,
        row_index=totals_row_index,
        col_index=6,
        h_align="LEFT",
    )
    story.append(table)

    if quote.note:
        story.extend([Spacer(1, 4 * mm), Paragraph(f"備註：{quote.note}", body_style)])
    story.extend([Spacer(1, 4 * mm), Paragraph("經手人：莊全立", signer_style)])

    doc.build(story, canvasmaker=_make_pdf_stamp_canvasmaker(doc, stamp_center))
    buffer.seek(0)
    return buffer


def _build_invoice_template_pdf(invoice: Invoice, customer: Customer | None, contact: Contact | None):
    _require_embedded_pdf_font()

    customer_name = (customer.name if customer else "") or ""
    contact_name = (contact.name if contact else "") or ""
    recipient = contact_name or customer_name
    if customer_name and contact_name and customer_name != contact_name:
        recipient = f"{customer_name} {contact_name}"

    ordered_items = sorted(
        invoice.items,
        key=lambda item: (
            item.sort_order if item.sort_order is not None else 10**9,
            item.id if item.id is not None else 10**9,
        ),
    )

    rows = [["項次", "項目名稱", "規格內容", "單位", "數量", "單價", "金額", "備註"]]
    for idx in range(20):
        item = ordered_items[idx] if idx < len(ordered_items) else None
        if item is None:
            rows.append([str(idx + 1), "", "", "", "", "", "", ""])
            continue
        rows.append(
            [
                str(idx + 1),
                item.description or "",
                "",
                item.unit or "",
                f"{float(item.quantity or 0):.2f}",
                f"{float(item.unit_price or 0):.2f}",
                f"{float(item.amount or 0):.2f}",
                "",
            ]
        )

    total_amount = _invoice_display_total_without_tax(invoice)
    rows.append(
        [
            "合計",
            "",
            "新台幣",
            f"{total_amount:.2f}",
            "",
            "NT$",
            f"{total_amount:.2f}",
            "",
        ]
    )

    issue_date_text = _to_roc_date_text(invoice.issue_date)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"invoice-{invoice.invoice_no}",
    )

    styles = getSampleStyleSheet()
    title_style = styles["Heading1"].clone("InvoiceTemplateTitle")
    title_style.fontName = PDF_FONT_NAME
    title_style.alignment = 1
    title_style.fontSize = 18
    title_style.leading = 22

    subtitle_style = styles["Normal"].clone("InvoiceTemplateSubtitle")
    subtitle_style.fontName = PDF_FONT_NAME
    subtitle_style.alignment = 1
    subtitle_style.fontSize = 12
    subtitle_style.leading = 16

    body_style = styles["Normal"].clone("InvoiceTemplateBody")
    body_style.fontName = PDF_FONT_NAME
    body_style.fontSize = 10
    body_style.leading = 14

    story = [
        Paragraph("立翔水電行", title_style),
        Paragraph("發票", subtitle_style),
        Spacer(1, 3 * mm),
        Paragraph(f"{recipient} 台照", body_style),
        Paragraph(issue_date_text, body_style),
        Paragraph(f"單號：{invoice.invoice_no}", body_style),
        Spacer(1, 4 * mm),
    ]

    table = Table(
        rows,
        colWidths=[12 * mm, 46 * mm, 28 * mm, 14 * mm, 14 * mm, 20 * mm, 20 * mm, 20 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (5, -1), "CENTER"),
                ("ALIGN", (6, 0), (6, -1), "RIGHT"),
                ("ALIGN", (7, 0), (7, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#9ca3af")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#fafafa")]),
                ("FONTNAME", (0, -1), (-1, -1), PDF_FONT_NAME),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2ff")),
                ("LINEABOVE", (0, -1), (-1, -1), 1.0, colors.HexColor("#4b5563")),
            ]
        )
    )
    totals_row_index = len(rows) - 1 if rows else 0
    stamp_center = _estimate_table_cell_center(
        doc,
        story,
        table,
        row_index=totals_row_index,
        col_index=6,
        h_align="LEFT",
    )
    story.append(table)

    if invoice.note:
        story.extend([Spacer(1, 4 * mm), Paragraph(f"備註：{invoice.note}", body_style)])

    doc.build(story, canvasmaker=_make_pdf_stamp_canvasmaker(doc, stamp_center))
    buffer.seek(0)
    return buffer


def _trim(value):
    return (value or "").strip()


def _append_note(original: str | None, extra: str) -> str:
    current = _trim(original)
    return f"{current}\n{extra}" if current else extra


def _find_or_create_customer(name: str, phone: str, email: str, address: str, note_line: str):
    customer = None
    if phone:
        customer = Customer.query.filter(Customer.phone == phone).first()
    if not customer and email:
        customer = Customer.query.filter(Customer.email == email).first()
    if not customer and name:
        customer = Customer.query.filter(Customer.name == name).first()

    if customer:
        if phone and not customer.phone:
            customer.phone = phone
        if email and not customer.email:
            customer.email = email
        if address and not customer.address:
            customer.address = address
        customer.note = _append_note(customer.note, note_line)
        return customer

    base_name = name or f"WebBooking-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    candidate = base_name
    suffix = 2
    while Customer.query.filter(Customer.name == candidate).first():
        candidate = f"{base_name}-{suffix}"
        suffix += 1

    customer = Customer(
        name=candidate,
        email=email or None,
        phone=phone or None,
        address=address or None,
        note=note_line,
        created_by_id=None,
    )
    db.session.add(customer)
    db.session.flush()
    return customer


def _find_or_create_contact(customer: Customer, name: str, phone: str, email: str, note_line: str):
    contact = None
    if email:
        contact = Contact.query.filter(
            Contact.customer_id == customer.id,
            Contact.email == email,
        ).first()
    if not contact and phone:
        contact = Contact.query.filter(
            Contact.customer_id == customer.id,
            Contact.phone == phone,
            Contact.name == name,
        ).first()

    if contact:
        if phone and not contact.phone:
            contact.phone = phone
        if email and not contact.email:
            contact.email = email
        contact.note = _append_note(contact.note, note_line)
        return contact

    is_primary = Contact.query.filter(Contact.customer_id == customer.id).count() == 0
    contact = Contact(
        customer_id=customer.id,
        name=name or customer.name,
        email=email or None,
        phone=phone or None,
        is_primary=is_primary,
        note=note_line,
    )
    db.session.add(contact)
    db.session.flush()
    return contact


@crm_bp.post("/public/bookings")
def create_public_booking():
    data = request.get_json(silent=True) or {}
    name = _trim(data.get("name"))
    phone = _trim(data.get("phone"))
    email = _trim(data.get("email"))
    service = _trim(data.get("service"))
    message = _trim(data.get("message"))
    address = _trim(data.get("address"))
    source_url = _trim(data.get("source_url")) or _trim(request.referrer)
    user_agent = _trim(request.headers.get("User-Agent"))
    client_ip = _trim((request.headers.get("X-Forwarded-For") or "").split(",")[0]) or _trim(request.remote_addr)

    if not name:
        return jsonify({"msg": "name is required"}), 400
    if not phone:
        return jsonify({"msg": "phone is required"}), 400
    if not service:
        return jsonify({"msg": "service is required"}), 400

    booking_note_parts = [
        "Source: website booking",
        f"Service: {service}",
    ]
    if message:
        booking_note_parts.append(f"Message: {message}")
    if source_url:
        booking_note_parts.append(f"Source URL: {source_url}")
    if client_ip:
        booking_note_parts.append(f"IP: {client_ip}")
    if user_agent:
        booking_note_parts.append(f"UA: {user_agent}")
    booking_note = "\n".join(booking_note_parts)

    customer = _find_or_create_customer(
        name=name,
        phone=phone,
        email=email,
        address=address,
        note_line=booking_note,
    )
    contact = _find_or_create_contact(
        customer=customer,
        name=name,
        phone=phone,
        email=email,
        note_line=booking_note,
    )
    db.session.commit()

    return (
        jsonify(
            {
                "msg": "booking received",
                "customer_id": customer.id,
                "contact_id": contact.id,
            }
        ),
        201,
    )


@crm_bp.get("/catalog-items")
@role_required(*READ_ROLES)
def list_catalog_items():
    q = (request.args.get("q") or "").strip()
    include_inactive = request.args.get("include_inactive", "false").strip().lower() == "true"

    query = ServiceCatalogItem.query.order_by(ServiceCatalogItem.updated_at.desc())
    if not include_inactive:
        query = query.filter(ServiceCatalogItem.is_active.is_(True))
    if q:
        pattern = f"%{q}%"
        query = query.filter(
            (ServiceCatalogItem.name.ilike(pattern))
            | (ServiceCatalogItem.category.ilike(pattern))
        )
    rows = query.limit(500).all()
    return jsonify([row.to_dict() for row in rows])


@crm_bp.post("/catalog-items")
@role_required(*WRITE_ROLES)
def create_catalog_item():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"msg": "name is required"}), 400

    unit = (data.get("unit") or "").strip() or "式"
    unit_price, unit_price_err = _parse_float(data.get("unit_price", 0), "unit_price", minimum=0)
    if unit_price_err:
        return unit_price_err

    exists = ServiceCatalogItem.query.filter(ServiceCatalogItem.name == name).first()
    if exists:
        return jsonify({"msg": "catalog item name already exists"}), 400

    item = ServiceCatalogItem(
        name=name,
        unit=unit,
        unit_price=unit_price or 0.0,
        category=(data.get("category") or "").strip() or None,
        note=(data.get("note") or "").strip() or None,
        is_active=bool(data.get("is_active", True)),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@crm_bp.put("/catalog-items/<int:item_id>")
@role_required(*WRITE_ROLES)
def update_catalog_item(item_id: int):
    item = ServiceCatalogItem.query.get_or_404(item_id)
    data = request.get_json() or {}

    if "name" in data:
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"msg": "name is required"}), 400
        duplicate = ServiceCatalogItem.query.filter(
            ServiceCatalogItem.name == name,
            ServiceCatalogItem.id != item.id,
        ).first()
        if duplicate:
            return jsonify({"msg": "catalog item name already exists"}), 400
        item.name = name

    if "unit" in data:
        item.unit = (data.get("unit") or "").strip() or "式"

    if "unit_price" in data:
        unit_price, unit_price_err = _parse_float(data.get("unit_price"), "unit_price", minimum=0)
        if unit_price_err:
            return unit_price_err
        item.unit_price = unit_price or 0.0

    if "category" in data:
        item.category = (data.get("category") or "").strip() or None
    if "note" in data:
        item.note = (data.get("note") or "").strip() or None
    if "is_active" in data:
        item.is_active = bool(data.get("is_active"))

    db.session.commit()
    return jsonify(item.to_dict())


@crm_bp.get("/customers/<int:customer_id>/service-history")
@role_required(*READ_ROLES)
def customer_service_history(customer_id: int):
    customer = Customer.query.get_or_404(customer_id)
    quote_limit = request.args.get("quote_limit", default=50, type=int) or 50
    quote_limit = max(1, min(quote_limit, 200))
    return jsonify(_serialize_customer_service_history(customer, quote_limit=quote_limit, invoice_limit=0))


@crm_bp.get("/service-history")
@role_required(*READ_ROLES)
def search_service_history():
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify({"msg": "q is required"}), 400

    pattern = f"%{q}%"
    customers = (
        Customer.query.filter(
            (Customer.name.ilike(pattern))
            | (Customer.phone.ilike(pattern))
            | (Customer.email.ilike(pattern))
        )
        .order_by(Customer.updated_at.desc())
        .limit(20)
        .all()
    )
    result = []
    for customer in customers:
        result.append(_serialize_customer_service_history(customer, quote_limit=10, invoice_limit=0))
    return jsonify(result)


@crm_bp.get("/customers")
@role_required(*READ_ROLES)
def list_customers():
    q = (request.args.get("q") or "").strip()
    query = Customer.query.order_by(Customer.updated_at.desc())
    if q:
        pattern = f"%{q}%"
        query = query.filter(Customer.name.ilike(pattern))
    customers = query.limit(200).all()
    return jsonify([item.to_dict() for item in customers])


@crm_bp.post("/customers")
@role_required(*WRITE_ROLES)
def create_customer():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"msg": "name is required"}), 400

    exists = Customer.query.filter(Customer.name == name).first()
    if exists:
        return jsonify({"msg": "Customer name already exists"}), 400

    customer = Customer(
        name=name,
        tax_id=(data.get("tax_id") or "").strip() or None,
        email=(data.get("email") or "").strip() or None,
        phone=(data.get("phone") or "").strip() or None,
        address=(data.get("address") or "").strip() or None,
        note=(data.get("note") or "").strip() or None,
        created_by_id=get_current_user_id(),
    )
    db.session.add(customer)
    db.session.commit()
    return jsonify(customer.to_dict()), 201


@crm_bp.get("/customers/<int:customer_id>")
@role_required(*READ_ROLES)
def get_customer(customer_id: int):
    customer = Customer.query.get_or_404(customer_id)
    payload = customer.to_dict()
    payload["contacts"] = [contact.to_dict() for contact in customer.contacts]
    return jsonify(payload)


@crm_bp.put("/customers/<int:customer_id>")
@role_required(*WRITE_ROLES)
def update_customer(customer_id: int):
    customer = Customer.query.get_or_404(customer_id)
    data = request.get_json() or {}

    if "name" in data:
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"msg": "name is required"}), 400
        duplicate = Customer.query.filter(Customer.name == name, Customer.id != customer_id).first()
        if duplicate:
            return jsonify({"msg": "Customer name already exists"}), 400
        customer.name = name

    for key in ("tax_id", "email", "phone", "address", "note"):
        if key in data:
            value = data.get(key)
            customer.__setattr__(key, (value or "").strip() or None)

    db.session.commit()
    return jsonify(customer.to_dict())


@crm_bp.get("/contacts")
@role_required(*READ_ROLES)
def list_contacts():
    customer_id = request.args.get("customer_id", type=int)
    query = Contact.query.order_by(Contact.updated_at.desc())
    if customer_id:
        query = query.filter(Contact.customer_id == customer_id)
    contacts = query.limit(300).all()
    return jsonify([item.to_dict() for item in contacts])


@crm_bp.post("/contacts")
@role_required(*WRITE_ROLES)
def create_contact():
    data = request.get_json() or {}
    customer_id = data.get("customer_id")
    name = (data.get("name") or "").strip()

    if not customer_id:
        return jsonify({"msg": "customer_id is required"}), 400
    if not name:
        return jsonify({"msg": "name is required"}), 400

    customer = Customer.query.get(customer_id)
    if not customer:
        return jsonify({"msg": "Customer not found"}), 404

    contact = Contact(
        customer_id=customer.id,
        name=name,
        title=(data.get("title") or "").strip() or None,
        email=(data.get("email") or "").strip() or None,
        phone=(data.get("phone") or "").strip() or None,
        is_primary=bool(data.get("is_primary")),
        note=(data.get("note") or "").strip() or None,
    )
    db.session.add(contact)
    db.session.commit()
    return jsonify(contact.to_dict()), 201


@crm_bp.put("/contacts/<int:contact_id>")
@role_required(*WRITE_ROLES)
def update_contact(contact_id: int):
    contact = Contact.query.get_or_404(contact_id)
    data = request.get_json() or {}

    if "customer_id" in data:
        next_customer = Customer.query.get(data.get("customer_id"))
        if not next_customer:
            return jsonify({"msg": "Customer not found"}), 404
        contact.customer_id = next_customer.id

    if "name" in data:
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"msg": "name is required"}), 400
        contact.name = name

    for key in ("title", "email", "phone", "note"):
        if key in data:
            value = data.get(key)
            contact.__setattr__(key, (value or "").strip() or None)

    if "is_primary" in data:
        contact.is_primary = bool(data.get("is_primary"))

    db.session.commit()
    return jsonify(contact.to_dict())


@crm_bp.get("/quotes")
@role_required(*READ_ROLES)
def list_quotes():
    query = Quote.query.options(selectinload(Quote.items)).order_by(Quote.updated_at.desc())
    customer_id = request.args.get("customer_id", type=int)
    status = (request.args.get("status") or "").strip().lower()

    if customer_id:
        query = query.filter(Quote.customer_id == customer_id)
    if status:
        query = query.filter(Quote.status == status)

    rows = query.limit(200).all()
    return jsonify([row.to_dict() for row in rows])


@crm_bp.post("/quotes")
@role_required(*WRITE_ROLES)
def create_quote():
    data = request.get_json() or {}
    customer_id = data.get("customer_id")
    contact_id = data.get("contact_id")
    status = (data.get("status") or "draft").strip().lower()
    if status not in VALID_QUOTE_STATUS:
        return jsonify({"msg": "Invalid quote status"}), 400

    if not customer_id:
        return jsonify({"msg": "customer_id is required"}), 400
    customer, contact, cc_err = _validate_customer_contact(customer_id, contact_id)
    if cc_err:
        return cc_err

    items, items_err = _normalize_items(data.get("items"))
    if items_err:
        return items_err

    issue_date, issue_err = _parse_date(data.get("issue_date"), "issue_date")
    if issue_err:
        return issue_err
    expiry_date, expiry_err = _parse_date(data.get("expiry_date"), "expiry_date")
    if expiry_err:
        return expiry_err
    issue_date, expiry_date = _default_quote_dates(issue_date, expiry_date)

    quote = Quote(
        quote_no=(data.get("quote_no") or "").strip() or _next_quote_no(),
        status=status,
        customer_id=customer_id,
        contact_id=contact_id,
        recipient_name=(data.get("recipient_name") or "").strip() or None,
        issue_date=issue_date,
        expiry_date=expiry_date,
        currency=(data.get("currency") or "TWD").strip().upper() or "TWD",
        note=(data.get("note") or "").strip() or None,
        created_by_id=get_current_user_id(),
    )

    total_err = _apply_totals(quote, items, data.get("tax_rate", 0))
    if total_err:
        return total_err

    db.session.add(quote)
    db.session.flush()

    for item in items:
        db.session.add(QuoteItem(quote_id=quote.id, **item))

    db.session.flush()
    quote = Quote.query.options(selectinload(Quote.items)).get(quote.id)
    if quote:
        _append_quote_version_snapshot(quote, action="create", summary="Initial quote created")
        db.session.add(_create_task_for_quote(quote, customer, contact))

    db.session.commit()
    quote = Quote.query.options(selectinload(Quote.items)).get(quote.id)
    return jsonify(quote.to_dict() if quote else {}), 201


@crm_bp.put("/quotes/<int:quote_id>")
@role_required(*WRITE_ROLES)
def update_quote(quote_id: int):
    quote = Quote.query.options(selectinload(Quote.items)).get_or_404(quote_id)
    data = request.get_json() or {}

    if "status" in data:
        status = (data.get("status") or "").strip().lower()
        if status not in VALID_QUOTE_STATUS:
            return jsonify({"msg": "Invalid quote status"}), 400
        quote.status = status

    next_customer_id = data.get("customer_id", quote.customer_id)
    next_contact_id = data.get("contact_id", quote.contact_id)
    _, _, cc_err = _validate_customer_contact(next_customer_id, next_contact_id)
    if cc_err:
        return cc_err
    quote.customer_id = next_customer_id
    quote.contact_id = next_contact_id

    if "issue_date" in data:
        parsed, err = _parse_date(data.get("issue_date"), "issue_date")
        if err:
            return err
        quote.issue_date = parsed

    if "expiry_date" in data:
        parsed, err = _parse_date(data.get("expiry_date"), "expiry_date")
        if err:
            return err
        quote.expiry_date = parsed

    if "currency" in data:
        quote.currency = (data.get("currency") or "TWD").strip().upper() or "TWD"
    if "recipient_name" in data:
        quote.recipient_name = (data.get("recipient_name") or "").strip() or None
    if "note" in data:
        quote.note = (data.get("note") or "").strip() or None

    if "items" in data:
        items, items_err = _normalize_items(data.get("items"))
        if items_err:
            return items_err
        quote.items.clear()
        db.session.flush()
        for item in items:
            db.session.add(QuoteItem(quote_id=quote.id, **item))
        total_err = _apply_totals(quote, items, data.get("tax_rate", quote.tax_rate))
    elif "tax_rate" in data:
        items = [item.to_dict() for item in quote.items]
        total_err = _apply_totals(quote, items, data.get("tax_rate", quote.tax_rate))
    else:
        total_err = None

    if total_err:
        return total_err

    db.session.flush()
    quote = Quote.query.options(selectinload(Quote.items)).get(quote.id)
    if quote:
        _append_quote_version_snapshot(quote, action="update", summary="Quote updated")

    db.session.commit()
    quote = Quote.query.options(selectinload(Quote.items)).get(quote.id)
    return jsonify(quote.to_dict())


@crm_bp.get("/quotes/<int:quote_id>/versions")
@role_required(*READ_ROLES)
def quote_versions(quote_id: int):
    quote = Quote.query.get_or_404(quote_id)
    rows = (
        QuoteVersion.query.options(selectinload(QuoteVersion.changed_by))
        .filter(QuoteVersion.quote_id == quote.id)
        .order_by(QuoteVersion.version_no.desc(), QuoteVersion.id.desc())
        .all()
    )
    return jsonify(
        {
            "quote_id": quote.id,
            "quote_no": quote.quote_no,
            "versions": [row.to_dict() for row in rows],
        }
    )


@crm_bp.post("/quotes/<int:quote_id>/convert-to-invoice")
@role_required(*WRITE_ROLES)
def convert_quote_to_invoice(quote_id: int):
    return _invoice_module_disabled()


@crm_bp.get("/invoices")
@role_required(*READ_ROLES)
def list_invoices():
    return _invoice_module_disabled()


@crm_bp.post("/invoices")
@role_required(*WRITE_ROLES)
def create_invoice():
    return _invoice_module_disabled()


@crm_bp.put("/invoices/<int:invoice_id>")
@role_required(*WRITE_ROLES)
def update_invoice(invoice_id: int):
    return _invoice_module_disabled()


@crm_bp.get("/quotes/<int:quote_id>/xlsx")
@role_required(*READ_ROLES)
def quote_xlsx(quote_id: int):
    quote = Quote.query.options(selectinload(Quote.items)).get_or_404(quote_id)
    customer = Customer.query.get(quote.customer_id)
    contact = Contact.query.get(quote.contact_id) if quote.contact_id else None

    template_path = _find_quote_template_path()
    if template_path is None:
        return jsonify({"msg": "Quote template not found"}), 500

    try:
        workbook = load_workbook(template_path)
    except Exception:
        return jsonify({"msg": "Quote template cannot be opened"}), 500

    if not workbook.worksheets:
        return jsonify({"msg": "Quote template has no worksheet"}), 500

    sheet_name = (request.args.get("sheet") or "").strip()
    if sheet_name:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if worksheet is None:
            return jsonify({"msg": f"Template sheet not found: {sheet_name}"}), 400
    else:
        worksheet = workbook.worksheets[0]

    for existing_sheet in list(workbook.worksheets):
        if existing_sheet.title != worksheet.title:
            workbook.remove(existing_sheet)

    safe_title = (quote.quote_no or "估價單").strip()[:31] or "估價單"
    worksheet.title = safe_title
    _apply_quote_to_template_sheet(worksheet, quote, customer, contact)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    customer_part = _safe_download_filename_part(customer.name if customer else None, fallback="客戶")
    quote_part = _safe_download_filename_part(quote.quote_no, fallback="估價單")
    filename = f"{customer_part}-{quote_part}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@crm_bp.get("/quotes/<int:quote_id>/pdf")
@role_required(*READ_ROLES)
def quote_pdf(quote_id: int):
    quote = Quote.query.options(selectinload(Quote.items)).get_or_404(quote_id)
    customer = Customer.query.get(quote.customer_id)
    contact = Contact.query.get(quote.contact_id) if quote.contact_id else None

    try:
        buffer = _build_quote_template_pdf(quote, customer, contact)
    except RuntimeError as exc:
        return jsonify(
            {
                "msg": "PDF 字型未就緒：目前未使用可嵌入的繁中字型，已停止輸出以避免手機/部分電腦顯示異常。",
                "detail": str(exc),
                "font_health": _pdf_font_health_payload(),
            }
        ), 500
    customer_raw = (customer.name if customer else None) or (quote.recipient_name or "").strip() or (contact.name if contact else None)
    customer_part = _safe_download_filename_part(customer_raw, fallback="客戶")
    if quote.issue_date:
        date_compact = quote.issue_date.strftime("%Y%m%d")
    elif getattr(quote, "created_at", None):
        date_compact = quote.created_at.strftime("%Y%m%d")
    else:
        date_compact = datetime.utcnow().strftime("%Y%m%d")
    date_part = _safe_download_filename_part(date_compact, fallback="date")
    filename = f"{customer_part}_{date_part}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=filename,
    )


@crm_bp.get("/health/pdf-font")
@role_required(*READ_ROLES)
def crm_pdf_font_health():
    return jsonify(_pdf_font_health_payload())


@crm_bp.get("/health/db")
@role_required(*READ_ROLES)
def crm_db_health():
    engine = db.engine
    url = engine.url
    return jsonify(
        {
            "dialect": engine.dialect.name,
            "driver": engine.dialect.driver,
            "database": url.database,
            "host": url.host,
            "port": url.port,
        }
    )


@crm_bp.get("/invoices/<int:invoice_id>/pdf")
@role_required(*READ_ROLES)
def invoice_pdf(invoice_id: int):
    return _invoice_module_disabled()


@crm_bp.get("/boot")
@jwt_required()
def crm_bootstrap():
    customers = Customer.query.order_by(Customer.updated_at.desc()).limit(50).all()
    contacts = Contact.query.order_by(Contact.updated_at.desc()).limit(100).all()
    quotes = Quote.query.options(selectinload(Quote.items)).order_by(Quote.updated_at.desc()).limit(30).all()
    catalog_items = (
        ServiceCatalogItem.query.filter(ServiceCatalogItem.is_active.is_(True))
        .order_by(ServiceCatalogItem.updated_at.desc())
        .limit(200)
        .all()
    )

    return jsonify(
        {
            "customers": [row.to_dict() for row in customers],
            "contacts": [row.to_dict() for row in contacts],
            "quotes": [row.to_dict() for row in quotes],
            "invoices": [],
            "catalog_items": [row.to_dict() for row in catalog_items],
        }
    )


